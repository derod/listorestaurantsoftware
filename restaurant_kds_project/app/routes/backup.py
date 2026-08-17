"""
Respaldo y Datos (admin).

- Respaldo completo del sistema: descarga un .zip con un snapshot consistente de
  la base de datos (SQLite online-backup) + la carpeta uploads/.
- Exportar / Importar catálogos (Productos e Insumos) en Excel y CSV. El import
  hace upsert por nombre (crea/actualiza, nunca borra) con vista previa.

Todo detrás del login de admin. El import valida y recalcula en el servidor.
"""
from __future__ import annotations

import csv
import io as _io
import sqlite3
import zipfile
from pathlib import Path

from fastapi import APIRouter, Depends, File, Form, Request, UploadFile
from fastapi.responses import JSONResponse, RedirectResponse, StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session

from ..database import get_db, DATA_DIR, engine
from ..models import Product, Ingredient, cr_now
from .web import templates, require_admin

router = APIRouter()


def _admin(request: Request) -> bool:
    return require_admin(request)


# ─── columnas de los catálogos (export e import usan las mismas) ──────────────

PRODUCT_COLUMNS = ["name", "category", "price", "active"]
INGREDIENT_COLUMNS = [
    "name", "category", "unit", "cost_per_unit", "stock",
    "purchase_unit", "pack_content", "purchase_price",
    "yield_qty", "yield_unit", "min_stock", "supplier", "status", "notes",
]


def _num(v, default=0.0):
    if v is None or v == "":
        return default
    try:
        return float(str(v).replace(",", "").strip())
    except (TypeError, ValueError):
        return default


def _bool(v, default=True):
    if v is None or v == "":
        return default
    s = str(v).strip().lower()
    if s in ("1", "true", "si", "sí", "yes", "y", "x", "activo"):
        return True
    if s in ("0", "false", "no", "n", "inactivo"):
        return False
    return default


def _str(v):
    return "" if v is None else str(v).strip()


# ─── respaldo completo (.zip) ─────────────────────────────────────────────────

def _sqlite_path() -> Path | None:
    try:
        if engine.url.get_backend_name() == "sqlite" and engine.url.database:
            p = Path(engine.url.database)
            return p if p.exists() else None
    except Exception:
        pass
    p = DATA_DIR / "restaurant_kds.db"
    return p if p.exists() else None


@router.get("/admin/backup")
def backup_page(request: Request, db: Session = Depends(get_db)):
    if not _admin(request):
        return RedirectResponse(url="/admin/login")
    stats = {
        "productos": db.query(func.count(Product.id)).scalar() or 0,
        "insumos": db.query(func.count(Ingredient.id)).scalar() or 0,
    }
    up = DATA_DIR / "uploads"
    n_files = sum(1 for f in up.rglob("*") if f.is_file()) if up.is_dir() else 0
    return templates.TemplateResponse(
        "admin_backup.html",
        {"request": request, "stats": stats, "n_files": n_files, "page_title": "Respaldo y Datos"},
    )


@router.get("/admin/backup/full.zip")
def backup_full(request: Request, db: Session = Depends(get_db)):
    if not _admin(request):
        return RedirectResponse(url="/admin/login")
    ts = cr_now().strftime("%Y%m%d-%H%M")
    buf = _io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        # BD: snapshot consistente con la API de backup de SQLite (aunque la app
        # esté escribiendo). Se hace a un archivo temporal en memoria vía :memory:.
        src_path = _sqlite_path()
        if src_path:
            mem = sqlite3.connect(":memory:")
            src = sqlite3.connect(str(src_path))
            try:
                src.backup(mem)  # copia consistente
                dump = "\n".join(mem.iterdump())
            finally:
                src.close()
            # Guardamos tanto el .db binario como un .sql de respaldo/portabilidad.
            tmpf = _io.BytesIO()
            # Reconstruir un .db desde memoria: usar backup a archivo temporal real.
            import tempfile
            with tempfile.NamedTemporaryFile(suffix=".db", delete=False) as tf:
                tmp_path = tf.name
            dst = sqlite3.connect(tmp_path)
            try:
                mem.backup(dst)
            finally:
                dst.close()
                mem.close()
            z.write(tmp_path, "restaurant_kds.db")
            z.writestr("restaurant_kds.sql", dump)
            try:
                Path(tmp_path).unlink()
            except OSError:
                pass
        # uploads/**
        up = DATA_DIR / "uploads"
        if up.is_dir():
            for f in up.rglob("*"):
                if f.is_file():
                    z.write(f, str(Path("uploads") / f.relative_to(up)))
        z.writestr("README.txt",
                   "Respaldo LISTO Restaurant Software\n"
                   f"Generado: {cr_now().strftime('%Y-%m-%d %H:%M')} (hora CR)\n\n"
                   "Contiene:\n"
                   "- restaurant_kds.db  (base de datos completa)\n"
                   "- restaurant_kds.sql (volcado SQL, por portabilidad)\n"
                   "- uploads/           (imágenes y documentos)\n\n"
                   "Para restaurar: reemplaza restaurant_kds.db y uploads/ en el\n"
                   "volumen de datos (DATA_DIR) y reinicia la aplicación.\n")
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename=listo-backup-{ts}.zip"},
    )


# ─── exportar catálogos ───────────────────────────────────────────────────────

def _rows_products(db: Session):
    yield PRODUCT_COLUMNS
    for p in db.query(Product).order_by(Product.display_order, Product.name).all():
        yield [p.name, p.category or "General", round(p.price or 0, 2), 1 if p.active else 0]


def _rows_ingredients(db: Session):
    yield INGREDIENT_COLUMNS
    for i in db.query(Ingredient).order_by(Ingredient.category, Ingredient.name).all():
        yield [
            i.name, i.category or "", i.unit or "", i.cost_per_unit or 0, i.stock or 0,
            i.purchase_unit or "", i.pack_content or "", i.purchase_price or "",
            i.yield_qty or "", i.yield_unit or "", i.min_stock or 0, i.supplier or "",
            i.status or "activo", i.notes or "",
        ]


def _csv_bytes(rows_iter) -> bytes:
    buf = _io.StringIO()
    w = csv.writer(buf)
    for row in rows_iter:
        w.writerow(row)
    return buf.getvalue().encode("utf-8-sig")


def _xlsx_bytes(rows_iter, sheet_title: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    rows = list(rows_iter)
    for r_i, row in enumerate(rows, start=1):
        for c_i, val in enumerate(row, start=1):
            ws.cell(row=r_i, column=c_i, value=val)
    if rows:
        for c_i in range(1, len(rows[0]) + 1):
            cell = ws.cell(row=1, column=c_i)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0F4C81")
            ws.column_dimensions[cell.column_letter].width = 18
    out = _io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _download(content: bytes, media: str, filename: str):
    return StreamingResponse(_io.BytesIO(content), media_type=media,
                             headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/admin/backup/export/{entity}.{fmt}")
def export_catalog(entity: str, fmt: str, request: Request, db: Session = Depends(get_db)):
    if not _admin(request):
        return RedirectResponse(url="/admin/login")
    if entity not in ("productos", "insumos") or fmt not in ("csv", "xlsx"):
        return RedirectResponse(url="/admin/backup")
    rows = _rows_products(db) if entity == "productos" else _rows_ingredients(db)
    ts = cr_now().strftime("%Y%m%d")
    if fmt == "csv":
        return _download(_csv_bytes(rows), "text/csv", f"{entity}-{ts}.csv")
    title = "Productos" if entity == "productos" else "Insumos"
    return _download(_xlsx_bytes(rows, title),
                     "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     f"{entity}-{ts}.xlsx")


# ─── importar catálogos (upsert por nombre) ───────────────────────────────────

def _read_rows(filename: str, content: bytes):
    ext = Path(filename or "").suffix.lower()
    if ext == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(_io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        raw = list(ws.iter_rows(values_only=True))
        if not raw:
            return []
        headers = [(_str(h).lower()) for h in raw[0]]
        out = []
        for r in raw[1:]:
            if r is None:
                continue
            d = {}
            for i, h in enumerate(headers):
                if h:
                    d[h] = r[i] if i < len(r) else None
            if any(v not in (None, "") for v in d.values()):
                out.append(d)
        return out
    # CSV
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(_io.StringIO(text))
    out = []
    for row in reader:
        d = {(_str(k).lower()): v for k, v in row.items() if k}
        if any(_str(v) for v in d.values()):
            out.append(d)
    return out


def _plan(db: Session, entity: str, rows: list[dict]):
    if entity == "productos":
        existing = {p.name.strip().lower(): p for p in db.query(Product).all()}
    else:
        existing = {i.name.strip().lower(): i for i in db.query(Ingredient).all()}
    plan = []
    for r in rows:
        name = _str(r.get("name") or r.get("nombre"))
        if not name:
            plan.append({"action": "omitir", "name": "(sin nombre)", "detail": "falta la columna name"})
            continue
        action = "actualizar" if name.lower() in existing else "crear"
        if entity == "productos":
            detail = f"{r.get('category') or 'General'} · ₡{round(_num(r.get('price')))}"
        else:
            detail = f"{_str(r.get('unit')) or 'unid'} · stock {_num(r.get('stock'))}"
        plan.append({"action": action, "name": name, "detail": detail, "raw": r})
    return plan


def _apply_plan(db: Session, entity: str, plan: list[dict]):
    created = updated = skipped = 0
    if entity == "productos":
        by_name = {p.name.strip().lower(): p for p in db.query(Product).all()}
        max_order = db.query(func.max(Product.display_order)).scalar() or 0
        for item in plan:
            if item["action"] == "omitir":
                skipped += 1
                continue
            r = item["raw"]
            name = item["name"]
            p = by_name.get(name.lower())
            if not p:
                max_order += 1
                p = Product(name=name[:200], display_order=max_order)
                db.add(p)
                by_name[name.lower()] = p
                created += 1
            else:
                updated += 1
            if _str(r.get("category")):
                p.category = _str(r.get("category"))[:40]
            if r.get("price") not in (None, ""):
                p.price = max(0.0, _num(r.get("price")))
            if r.get("active") not in (None, ""):
                p.active = _bool(r.get("active"), p.active)
    else:
        by_name = {i.name.strip().lower(): i for i in db.query(Ingredient).all()}
        for item in plan:
            if item["action"] == "omitir":
                skipped += 1
                continue
            r = item["raw"]
            name = item["name"]
            ing = by_name.get(name.lower())
            if not ing:
                ing = Ingredient(name=name[:200])
                db.add(ing)
                by_name[name.lower()] = ing
                created += 1
            else:
                updated += 1
            for col, attr in [("unit", "unit"), ("purchase_unit", "purchase_unit"),
                              ("yield_unit", "yield_unit"), ("supplier", "supplier"),
                              ("status", "status"), ("category", "category"), ("notes", "notes")]:
                if _str(r.get(col)):
                    setattr(ing, attr, _str(r.get(col))[:200])
            for col, attr in [("stock", "stock"), ("min_stock", "min_stock"),
                              ("pack_content", "pack_content"), ("purchase_price", "purchase_price"),
                              ("yield_qty", "yield_qty"), ("cost_per_unit", "cost_per_unit")]:
                if r.get(col) not in (None, ""):
                    setattr(ing, attr, _num(r.get(col)))
            # Recalcular costo por unidad base si hay presentación de compra.
            if ing.purchase_price and ing.pack_content and ing.pack_content > 0:
                ing.cost_per_unit = round(ing.purchase_price / ing.pack_content, 4)
    db.commit()
    return {"created": created, "updated": updated, "skipped": skipped}


@router.post("/admin/backup/import/{mode}")
async def import_catalog(mode: str, request: Request,
                         entity: str = Form(...), file: UploadFile = File(...),
                         db: Session = Depends(get_db)):
    if not _admin(request):
        return JSONResponse({"error": "No autorizado"}, status_code=401)
    if entity not in ("productos", "insumos") or mode not in ("preview", "apply"):
        return JSONResponse({"error": "Parámetros inválidos"}, status_code=400)
    content = await file.read()
    try:
        rows = _read_rows(file.filename, content)
    except Exception as exc:
        return JSONResponse({"error": f"No se pudo leer el archivo: {exc}"}, status_code=400)
    if not rows:
        return JSONResponse({"error": "El archivo no tiene filas de datos."}, status_code=400)
    plan = _plan(db, entity, rows)
    crear = sum(1 for p in plan if p["action"] == "crear")
    actualizar = sum(1 for p in plan if p["action"] == "actualizar")
    omitir = sum(1 for p in plan if p["action"] == "omitir")
    if mode == "preview":
        return {
            "entity": entity, "total": len(plan),
            "crear": crear, "actualizar": actualizar, "omitir": omitir,
            "rows": [{"action": p["action"], "name": p["name"], "detail": p.get("detail", "")} for p in plan[:300]],
            "truncated": len(plan) > 300,
        }
    result = _apply_plan(db, entity, plan)
    return {"ok": True, **result}
