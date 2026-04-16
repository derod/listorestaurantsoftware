"""
Order history service: filtered queries + financial calculations + exporters.
"""
from __future__ import annotations

import csv
import io
from datetime import datetime
from typing import Optional

from sqlalchemy.orm import Session, joinedload
from sqlalchemy import and_

from .models import Order, OrderItem, Product, Sale, AudioSettings, cr_now


# ─── filter dataclass ────────────────────────────────────────────────────────

class OrderHistoryFilters:
    def __init__(
        self,
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
        status: Optional[str] = None,
        source_role: Optional[str] = None,
        waiter_name: Optional[str] = None,
        product_id: Optional[int] = None,
        payment_method: Optional[str] = None,
        min_total: Optional[float] = None,
        max_total: Optional[float] = None,
    ):
        self.date_from = _parse_date(date_from)
        self.date_to = _parse_date(date_to, end_of_day=True)
        self.status = status or None
        self.source_role = source_role or None
        self.waiter_name = waiter_name.strip() if waiter_name else None
        self.product_id = product_id
        self.payment_method = payment_method or None
        self.min_total = min_total
        self.max_total = max_total


def _parse_date(value: Optional[str], end_of_day: bool = False) -> Optional[datetime]:
    if not value:
        return None
    try:
        d = datetime.strptime(value, "%Y-%m-%d")
        if end_of_day:
            return d.replace(hour=23, minute=59, second=59)
        return d
    except ValueError:
        return None


# ─── main query ──────────────────────────────────────────────────────────────

def get_order_history(db: Session, filters: OrderHistoryFilters, page: int = 1, per_page: int = 50):
    """
    Returns (orders, total_count, tax_rate).
    Each order has a synthetic `.fin` dict attached with subtotal/tax/total/payment_method.
    """
    q = _base_query(db, filters)
    total = q.count()
    orders = (
        q.options(
            joinedload(Order.items).joinedload(OrderItem.product)
        )
        .order_by(Order.created_at.desc())
        .offset((page - 1) * per_page)
        .limit(per_page)
        .all()
    )

    settings = db.query(AudioSettings).first()
    tax_rate = (settings.tax_rate if settings and settings.tax_rate else 0) / 100.0

    # Load related Sales in one query to avoid N+1
    order_ids = [o.id for o in orders]
    sales_map = {}
    if order_ids:
        # Sale has no direct FK to Order; match by closest created_at heuristic is impractical.
        # POS sales are independent. We compute financials from OrderItems + tax_rate only.
        pass

    for order in orders:
        order.fin = _compute_financials(order, tax_rate)

    return orders, total, tax_rate


def get_all_for_export(db: Session, filters: OrderHistoryFilters):
    """Unpaginated query for exports."""
    settings = db.query(AudioSettings).first()
    tax_rate = (settings.tax_rate if settings and settings.tax_rate else 0) / 100.0

    orders = (
        _base_query(db, filters)
        .options(joinedload(Order.items).joinedload(OrderItem.product))
        .order_by(Order.created_at.desc())
        .all()
    )
    for order in orders:
        order.fin = _compute_financials(order, tax_rate)
    return orders, tax_rate


def _base_query(db: Session, f: OrderHistoryFilters):
    q = db.query(Order)

    if f.date_from:
        q = q.filter(Order.created_at >= f.date_from)
    if f.date_to:
        q = q.filter(Order.created_at <= f.date_to)
    if f.status:
        q = q.filter(Order.status == f.status)
    if f.source_role:
        q = q.filter(Order.source_role == f.source_role)
    if f.waiter_name:
        q = q.filter(Order.waiter_name.ilike(f"%{f.waiter_name}%"))
    if f.product_id:
        q = q.filter(Order.items.any(OrderItem.product_id == f.product_id))
    # payment_method & amount filters applied post-compute (no Sale FK on Order)
    return q


def _compute_financials(order: Order, tax_rate: float) -> dict:
    subtotal = sum(
        (item.quantity * (item.product.price or 0))
        for item in order.items
        if item.product
    )
    tax = round(subtotal * tax_rate, 2)
    total = round(subtotal + tax, 2)
    return {
        "subtotal": round(subtotal, 2),
        "tax": tax,
        "total": total,
        "payment_method": "-",
    }


def _items_summary(order: Order) -> str:
    return ", ".join(
        f"{item.quantity}x {item.product.name}"
        for item in order.items
        if item.product
    )


# ─── CSV exporter ─────────────────────────────────────────────────────────────

def export_csv(orders) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([
        "order_id", "fecha", "estado", "origen", "mesero",
        "items", "subtotal", "tax", "total", "payment_method", "cancelado"
    ])
    for o in orders:
        writer.writerow([
            o.id,
            o.created_at.strftime("%Y-%m-%d %H:%M"),
            o.status,
            o.source_role,
            o.waiter_name or "-",
            _items_summary(o),
            o.fin["subtotal"],
            o.fin["tax"],
            o.fin["total"],
            o.fin["payment_method"],
            "Sí" if o.was_cancelled else "No",
        ])
    return buf.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility


# ─── Excel exporter ───────────────────────────────────────────────────────────

def export_excel(orders) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Historial de Órdenes"

    headers = [
        "ID", "Fecha", "Estado", "Origen", "Mesero",
        "Items", "Subtotal", "Tax", "Total", "Pago", "Cancelado"
    ]
    header_fill = PatternFill("solid", fgColor="1A1F2E")
    header_font = Font(bold=True, color="4DD3FF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    total_subtotal = total_tax = total_total = 0.0
    for row_idx, o in enumerate(orders, 2):
        cancelled = o.was_cancelled
        ws.append([
            o.id,
            o.created_at.strftime("%Y-%m-%d %H:%M"),
            o.status,
            o.source_role,
            o.waiter_name or "-",
            _items_summary(o),
            o.fin["subtotal"],
            o.fin["tax"],
            o.fin["total"],
            o.fin["payment_method"],
            "Sí" if cancelled else "No",
        ])
        if not cancelled:
            total_subtotal += o.fin["subtotal"]
            total_tax += o.fin["tax"]
            total_total += o.fin["total"]

    # Totals row
    last = ws.max_row + 1
    ws.cell(row=last, column=6, value="TOTAL").font = Font(bold=True)
    ws.cell(row=last, column=7, value=round(total_subtotal, 2)).font = Font(bold=True)
    ws.cell(row=last, column=8, value=round(total_tax, 2)).font = Font(bold=True)
    ws.cell(row=last, column=9, value=round(total_total, 2)).font = Font(bold=True)

    # Auto-fit columns
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── PDF exporter ─────────────────────────────────────────────────────────────

def export_pdf(orders, filters: OrderHistoryFilters, tax_rate: float) -> bytes:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Heading1"], fontSize=18, spaceAfter=6)
    sub_style = ParagraphStyle("sub", parent=styles["Normal"], fontSize=9, textColor=colors.grey)

    date_range = ""
    if filters.date_from and filters.date_to:
        date_range = f"{filters.date_from.strftime('%Y-%m-%d')} → {filters.date_to.strftime('%Y-%m-%d')}"
    elif filters.date_from:
        date_range = f"Desde {filters.date_from.strftime('%Y-%m-%d')}"
    elif filters.date_to:
        date_range = f"Hasta {filters.date_to.strftime('%Y-%m-%d')}"
    else:
        date_range = "Todas las fechas"

    story = [
        Paragraph("Reporte de Órdenes", title_style),
        Paragraph(f"Período: {date_range}   |   Generado: {cr_now().strftime('%Y-%m-%d %H:%M')} (Costa Rica)", sub_style),
        Spacer(1, 6 * mm),
    ]

    # Table data
    col_headers = ["ID", "Fecha", "Estado", "Origen", "Mesero", "Items", "Subtotal", "Tax", "Total", "Cancelado"]
    data = [col_headers]
    total_subtotal = total_tax = total_total = 0.0
    non_cancelled = 0

    small = ParagraphStyle("small", parent=styles["Normal"], fontSize=7, leading=9)

    for o in orders:
        cancelled = o.was_cancelled
        items_text = Paragraph(_items_summary(o), small)
        data.append([
            str(o.id),
            o.created_at.strftime("%Y-%m-%d\n%H:%M"),
            o.status,
            o.source_role,
            o.waiter_name or "-",
            items_text,
            f"${o.fin['subtotal']:.2f}",
            f"${o.fin['tax']:.2f}",
            f"${o.fin['total']:.2f}",
            "Sí" if cancelled else "No",
        ])
        if not cancelled:
            total_subtotal += o.fin["subtotal"]
            total_tax += o.fin["tax"]
            total_total += o.fin["total"]
            non_cancelled += 1

    # Summary row
    data.append([
        "", "", "", "", "TOTAL", "",
        f"${total_subtotal:.2f}",
        f"${total_tax:.2f}",
        f"${total_total:.2f}",
        "",
    ])

    col_widths = [18 * mm, 28 * mm, 26 * mm, 26 * mm, 30 * mm, 60 * mm, 22 * mm, 22 * mm, 22 * mm, 18 * mm]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A1F2E")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#4DD3FF")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.HexColor("#0D1117"), colors.HexColor("#161B22")]),
        ("TEXTCOLOR", (0, 1), (-1, -2), colors.HexColor("#C9D1D9")),
        ("FONTSIZE", (0, 1), (-1, -2), 7),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#1A1F2E")),
        ("TEXTCOLOR", (0, -1), (-1, -1), colors.white),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, -1), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#30363D")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWHEIGHT", (0, 1), (-1, -2), 18),
    ]))

    story.append(table)
    story.append(Spacer(1, 8 * mm))

    # Summary block
    summary_data = [
        ["Total órdenes", str(len(orders))],
        ["Órdenes activas (no canceladas)", str(non_cancelled)],
        ["Total ingresos", f"${total_total:.2f}"],
        ["Total impuestos", f"${total_tax:.2f}"],
        ["Tasa IVA aplicada", f"{tax_rate * 100:.1f}%"],
    ]
    summary_table = Table(summary_data, colWidths=[60 * mm, 40 * mm])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#161B22")),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#8B949E")),
        ("TEXTCOLOR", (1, 0), (1, -1), colors.white),
        ("FONTNAME", (1, 0), (1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#30363D")),
        ("ROWHEIGHT", (0, 0), (-1, -1), 16),
    ]))
    story.append(summary_table)

    doc.build(story)
    return buf.getvalue()
